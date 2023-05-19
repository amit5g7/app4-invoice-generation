import pandas as pd
#import os
import openpyxl
import glob
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    df = pd.read_excel(filepath,sheet_name="Sheet 1")
    total_price_sum = df["total_price"].sum()
    #print(total_price_sum)
    #print(df)
    pdf= FPDF(orientation="P", unit= "mm", format="A4")
    pdf.add_page()
    filename=Path(filepath).stem
    invoice_nr = filename.split("-")[0]
    date_invoice=filename.split("-")[1]
    pdf.set_font(family="Times", size=14, style="B")
    pdf.cell(w=50,h=8,txt=f"Invoice No.{invoice_nr}",ln=1)
    pdf.cell(w=50,h=8,txt=f"Date: {date_invoice}",ln=1)
    pdf.ln(8)
    pdf.set_font(family="Times", size=12, style="B")
    for col in df.columns:
        pdf.cell(w=38, h=5, txt=f"{col.title().replace('_',' ')}", border=1,align="C")
    #pdf.ln(5)
    pdf.set_font(family="Times", size=8)
    for index, rows in df.iterrows():
        pdf.ln(5)
        for col in df.columns:
            pdf.cell(w=38, h=5, txt=f"{rows[col]}", border=1,align="C")
    pdf.ln(5)

    for col in df.columns:

        #print(total_price_sum)
        #print(col)
        if col == "total_price":
            #print(rows[col])
            pdf.cell(w=38, h=5, txt=f"{total_price_sum}", border=1, align="C",)
        else:
            pdf.cell(w=38, h=5, border=1, align="C", )

    pdf.ln(16)
    pdf.set_font(family="Times", size=12, style="B")
    pdf.cell(w=38, h=5, txt=f"The total due amount is {total_price_sum} Euros.", border=0, align="L",)





pdf.output(f"PDFs/{filename}.pdf")
